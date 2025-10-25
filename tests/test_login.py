# tests/test_login.py
import re, os, pathlib, pytest
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from pages.login_page import LoginPage

n = lambda s: re.sub(r"\s+", " ", (s or "")).strip().lower()
DATA_FILE = os.path.join(pathlib.Path(__file__).resolve().parents[1], "data", "login_test.xlsx")

def read_rows(xlsx):
    ws = load_workbook(xlsx)["Sheet1"]
    head = [str(h or "").strip().lower() for h in next(ws.iter_rows(values_only=True))]
    idx = {k: i for i, k in enumerate(head)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r): continue
        rows.append({"username": r[idx["username"]], "password": r[idx["password"]], "expected": r[idx["expected"]]})
    return rows

DATA = read_rows(DATA_FILE)

@pytest.mark.parametrize("row", DATA, ids=[f"row{i+1}" for i, _ in enumerate(DATA)])
def test_login(driver, base_url, row):
    email, pwd, exp = (str(row[k] or "") for k in ("username", "password", "expected"))
    page = LoginPage(driver)
    page.open(base_url)
    page.login(email, pwd)

    print("\nTEST LOGIN")
    print("Email:", email)
    print("Password:", pwd)
    print("Expected:", exp)

    if exp.startswith("http"):
        ok = WebDriverWait(driver, 12).until(lambda d: d.current_url.startswith(exp))
        print("Actual URL:", driver.current_url)
        print("Result:", "PASS" if ok else "FAIL")
        assert ok
    else:
        actual = page.get_error_text()
        print("Actual message:", actual)
        passed = n(exp) in n(actual)
        print("Result:", "PASS" if passed else "FAIL")
        assert passed, f"Expected: {exp}\nActual: {actual}"
