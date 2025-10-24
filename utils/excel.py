# E:\DoAn4\utils\excel.py
from openpyxl import load_workbook

def read_login_rows(xlsx_path: str, sheet="Sheet1"):
    """
    Đọc file Excel có cột: username, password, expected
    Trả về: list[dict] mỗi dòng là 1 test row (dùng cho login)
    """
    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' không tồn tại. Sheets: {wb.sheetnames}")
    ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    # Chuẩn hóa header để không lệ thuộc viết hoa/thường
    norm = [h.lower() for h in headers]

    # Tối thiểu phải có 3 cột
    if not all(col in norm for col in ["username", "password", "expected"]):
        raise ValueError(f"Các cột bắt buộc: username, password, expected. Thực tế: {headers}")

    data = []
    for r in rows[1:]:
        if not r or all(v is None for v in r):
            continue
        rowdict = dict(zip(headers, r))
        data.append({
            "username": rowdict.get("username"),
            "password": rowdict.get("password"),
            "expected": rowdict.get("expected"),
        })
    return data


def read_sheet(xlsx_path: str, sheet="Sheet1"):
    """
    Hàm đa năng dùng cho Đăng ký hoặc các test khác.
    - Đọc sheet Excel thành list[dict]
    - Header ở dòng đầu tiên
    - Tự chuẩn hóa key: về lowercase, thay khoảng trắng bằng _
    """
    wb = load_workbook(xlsx_path)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' không tồn tại. Sheets: {wb.sheetnames}")
    ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for r in rows[1:]:
        if not r or all(v is None for v in r):
            continue
        rowdict = dict(zip(headers, r))
        # Chuẩn hóa key: lowercase + thay khoảng trắng bằng "_"
        norm = {str(k).strip().lower().replace(" ", "_"): v for k, v in rowdict.items()}
        data.append(norm)
    return data
